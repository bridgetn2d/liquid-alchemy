#!/usr/bin/env python3
"""Import Attaboy Bug Cocktail Family entries into the canonical workbook."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

WORKBOOK = Path(__file__).resolve().parent.parent / (
    "content/workbook/Liquid_Alchemy_Content_Database_v1_repo_ready.xlsx"
)

PACK_ID = "pack_002_attaboy_bug_family"
PACK_NAME = "Attaboy Bug Cocktail Family"
SOURCE_URL = "https://punchdrink.com/articles/attaboy-equal-parts-cocktail-recipes/"
SOURCE_NOTES = (
    "Documented Attaboy specs via Punch (January 2025). "
    "Liquid Alchemy editorial framing; creator attribution preserved."
)

CRAFT_ITEM_ID = "sweetened-ginger-juice"
CRAFT_GINGER_TEXT = (
    "Juice fresh ginger skin-on, then mix 4 parts ginger juice to 3 parts "
    "granulated sugar by weight, stirring until the sugar fully dissolves. "
    "Keeps refrigerated for up to 2 days. Do not substitute bottled ginger "
    "juice — the fresh-pressed intensity is the point."
)

CREATORS = {
    "sam_ross_attaboy": {
        "name": "Sam Ross, Attaboy, New York City",
        "bio": "Sam Ross created the Mosquito at Attaboy (New York City) in 2017.",
        "source_urls": SOURCE_URL,
    },
    "brandon_bramhall_attaboy": {
        "name": "Brandon Bramhall, Attaboy, Nashville",
        "bio": "Brandon Bramhall created the Praying Mantis at Attaboy Nashville.",
        "source_urls": SOURCE_URL,
    },
    "mike_mccollum_attaboy": {
        "name": "Mike McCollum, Attaboy, Nashville",
        "bio": "Mike McCollum created the Maverick at Attaboy Nashville.",
        "source_urls": SOURCE_URL,
    },
    "jojo_colona_attaboy": {
        "name": "Jojo Colona, Attaboy, New York City",
        "bio": "Jojo Colona created the Lantern Fly at Attaboy New York City.",
        "source_urls": SOURCE_URL,
    },
}

NEW_INGREDIENTS = [
    ("sweetened_ginger_juice", "Sweetened ginger juice", "juice", "House Attaboy prep; see craft item sweetened-ginger-juice."),
    ("islay_scotch", "Islay Scotch", "spirit", None),
    ("gran_classico", "Gran Classico", "liqueur", None),
    ("amargo_vallet", "Amargo-Vallet amaro", "amaro", None),
    ("doctor_bird_jamaican_pot_still_rum", "Jamaican pot still rum (Doctor Bird recommended)", "spirit", None),
    ("candied_ginger", "Candied ginger", "garnish", None),
]

COCKTAILS = [
    {
        "recipe_id": "mosquito",
        "name": "Mosquito",
        "creator_id": "sam_ross_attaboy",
        "origin": "Sam Ross, Attaboy, New York City, 2017",
        "base_spirit": "Mezcal",
        "tags": "Sour/Tart, Bitter, Smoky, Spicy, Boozy",
        "sliders": dict(boozy=6, sweet=3, sour=6, bitter=7, fruity=2, herbal=2, smoky=6, spicy=5, rich=1),
        "related_recipe_ids": "paper_plane,last_word",
        "ingredients": [
            ("mezcal", "Mezcal", "0.75", "oz"),
            ("fresh_lemon_juice", "Fresh lemon juice", "0.75", "oz"),
            ("sweetened_ginger_juice", "Sweetened ginger juice", "0.75", "oz"),
            ("campari", "Campari", "0.75", "oz"),
        ],
        "instruction": (
            "Combine all ingredients in a cocktail shaker with ice. Shake until well chilled. "
            "Strain into a chilled coupe. Garnish with candied ginger."
        ),
        "garnish": ("Candied ginger", None, "coupe rim or pick"),
        "story": """Sam Ross created the Paper Plane in 2007 for the opening menu of Chicago's Violet Hour, drawing on the Last Word — the pre-Prohibition equal-parts classic of gin, green Chartreuse, maraschino, and lime — as his structural template. Equal parts, shaken, citrus-forward, bitter element. The Paper Plane went on to become one of the most widely traveled modern classics in the craft cocktail canon.

Ten years later, Ross revisited his own creation. At Attaboy — the celebrated bar he launched on Manhattan's Lower East Side in 2012 with partner Michael McIlroy — he assembled a new equal-parts drink called the Mosquito: mezcal, lemon, sweetened ginger juice, Campari. The architecture is the same as the Paper Plane. The personality is entirely different — smokier, spicier, more bitter, with ginger replacing Amaro Nonino's herbal sweetness and Campari sharpening the finish where Aperol had softened it.

The Mosquito became the progenitor of what Attaboy staff have dubbed the Bug Cocktail family: a growing library of equal-parts originals, now numbering more than eight, each keeping the lemon and ginger components intact while swapping the spirit and bitter element. It is, in miniature, the same creative logic that produced the Mosquito from the Paper Plane — a template passed forward, remixed, extended. Attaboy operates without written menus, relying on bartender's choice, which means these drinks exist in the bar's institutional memory as much as on any page. Publishing them is an act of preservation as much as documentation.""",
        "notes": """Creator's spec per Punch, January 2025. The ginger juice is the heart of this drink and every drink in the Bug family — it must be made fresh. The sweetening ratio (4:3 juice to sugar by weight) is precise and intentional; adjust and the balance collapses. Campari's bitterness is what reins in the ginger's natural sweetness — they are calibrated against each other. Mezcal with genuine smoke presence is necessary; a neutral-tasting mezcal loses the tension that makes the drink interesting.""",
        "riffs": """The Paper Plane (Sam Ross, 2007) is the direct ancestor: bourbon, Aperol, Amaro Nonino, lemon, equal parts. The Last Word (pre-Prohibition) is the architectural grandfather: gin, green Chartreuse, maraschino, lime, equal parts. The three Bug Cocktails below are all Mosquito variations following the same equal-parts ginger-lemon template with different spirit and bitter pairings.""",
        "why_it_matters": "The Mosquito is the seed of a creative system — a template remixed from a template, now producing its own family of drinks at one of the most influential bars in contemporary craft cocktail culture. The equal-parts format, descended from the Last Word through the Paper Plane, is one of the most generative structures in modern mixology.",
        "curiosity_hook": "Sam Ross has described the equal-parts format as a balancing discipline: every component must earn its quarter of the glass. There is nowhere to hide. A weak spirit, a stale citrus, an under-sweetened ginger juice — the format exposes all of it. This is the equal-parts lesson: simplicity is the hardest standard to meet.",
        "source_urls": "https://punchdrink.com/recipes/mosquito/\n" + SOURCE_URL,
    },
    {
        "recipe_id": "praying_mantis",
        "name": "Praying Mantis",
        "creator_id": "brandon_bramhall_attaboy",
        "origin": "Brandon Bramhall, Attaboy, Nashville",
        "base_spirit": "Scotch",
        "tags": "Sour/Tart, Bitter, Smoky, Boozy, Herbal/Botanical",
        "sliders": dict(boozy=7, sweet=2, sour=6, bitter=8, fruity=1, herbal=5, smoky=7, spicy=3, rich=1),
        "related_recipe_ids": "mosquito,paper_plane,last_word",
        "ingredients": [
            ("islay_scotch", "Islay Scotch", "0.75", "oz"),
            ("fresh_lemon_juice", "Fresh lemon juice", "0.75", "oz"),
            ("sweetened_ginger_juice", "Sweetened ginger juice", "0.75", "oz"),
            ("gran_classico", "Gran Classico", "0.75", "oz"),
        ],
        "instruction": (
            "Combine all ingredients in a cocktail shaker with ice. Shake until well chilled. "
            "Strain into a chilled coupe. Garnish with candied ginger."
        ),
        "garnish": ("Candied ginger", None, "coupe rim or pick"),
        "story": """Brandon Bramhall created the Praying Mantis at Attaboy's Nashville outpost, working within the equal-parts ginger-lemon template that Sam Ross established with the Mosquito. The formula is consistent across the Bug family: a base spirit, lemon juice, sweetened ginger juice, and a bitter liqueur or amaro, all in equal proportion, all shaken and served up in a coupe. What changes is the pairing — and the pairing is everything.

Bramhall's choice of Islay Scotch and Gran Classico is the most austere and complex variation in the Bug family. Islay Scotch brings peat smoke, maritime salt, and iodine — a flavor profile that has no equivalent in any other spirit and that pairs with bitterness in ways that require careful calibration. Gran Classico, the Swiss bitter liqueur made from a nineteenth-century recipe with twenty-five herbs and roots, sits between Campari and Aperol in sweetness but carries a complexity that neither achieves. The combination produces a drink that is simultaneously more smoky, more bitter, and more herbal than the Mosquito — and considerably more demanding of the drinker.

The name is well-chosen. The praying mantis is a patient, precise, and formidable creature. So is this drink.""",
        "notes": """Creator's spec per Punch, January 2025. Islay Scotch is not interchangeable with blended or Highland Scotch here — the peat is structural, not decorative. Laphroaig 10, Ardbeg 10, or Caol Ila 12 are all appropriate choices at different smoke intensities. Gran Classico is available in well-stocked craft spirits retailers; Contratto Bitter or Tempus Fugit Gran Classico are the same product. Do not substitute Campari — the bitterness profile is different enough to produce a meaningfully different drink.""",
        "riffs": """Substitute blended Scotch for Islay to soften the smoke and produce a more approachable version at some cost to complexity. Campari in place of Gran Classico sharpens the bitterness and reduces the herbal depth — closer to a Scotch Mosquito than a distinct drink. For a lower-proof version, substitute a peated Scotch-style whisky liqueur (Drambuie adds honey sweetness that changes the balance significantly — account for it by reducing the ginger juice).""",
        "why_it_matters": "The most complex and demanding drink in the Bug family — a demonstration that the equal-parts template can carry serious weight when the ingredients are chosen with precision. Islay Scotch in a shaken equal-parts sour is an unusual move that works.",
        "curiosity_hook": "Gran Classico and Campari occupy similar roles — bitter liqueur, similar color, similar ABV — but produce meaningfully different drinks. Understanding the functional differences between bitter liqueurs (sweetness level, herbal character, proof, finish) is essential to working intelligently with the Daisy template. Swapping one bitter for another is not a neutral substitution.",
        "source_urls": SOURCE_URL,
    },
    {
        "recipe_id": "maverick",
        "name": "Maverick",
        "creator_id": "mike_mccollum_attaboy",
        "origin": "Mike McCollum, Attaboy, Nashville",
        "base_spirit": "Gin",
        "tags": "Sour/Tart, Herbal/Botanical, Light/Refreshing, Fruity, Elegant",
        "sliders": dict(boozy=5, sweet=4, sour=6, bitter=4, fruity=2, herbal=6, smoky=0, spicy=3, rich=1),
        "related_recipe_ids": "mosquito,last_word,paper_plane",
        "ingredients": [
            ("london_dry_gin", "London dry gin", "0.75", "oz"),
            ("fresh_lemon_juice", "Fresh lemon juice", "0.75", "oz"),
            ("sweetened_ginger_juice", "Sweetened ginger juice", "0.75", "oz"),
            ("aperol", "Aperol", "0.75", "oz"),
            ("fresh_mint_leaves", "Fresh mint leaves", "7", "whole", "6–8 leaves; shaken in, not muddled"),
        ],
        "instruction": (
            "Combine all ingredients in a cocktail shaker with ice. Shake until well chilled. "
            "Strain through a tightly gated Hawthorne strainer — or double-strain — into a chilled coupe. "
            "The mint is shaken in, not muddled; the strainer catches the leaves. Garnish with candied ginger."
        ),
        "garnish": ("Candied ginger", None, "coupe rim or pick"),
        "story": """Mike McCollum of Attaboy Nashville brought two things to the Bug family that none of the other drinks in the series have: gin as the base spirit, and fresh mint shaken directly into the cocktail.

The gin choice is interesting. The Last Word — the great-grandparent of this entire family — is a gin drink. The Paper Plane moved to bourbon. The Mosquito moved to mezcal. The Maverick returns to gin, completing a kind of circle, and pairs it with Aperol rather than Campari or Gran Classico. The result is the lightest, most herbaceous, and most approachable drink in the Bug family — Aperol's lower bitterness and higher sweetness softening the ginger's heat, gin's botanicals amplified by the mint, lemon keeping everything bright and clean.

The mint technique is the detail worth noting. Rather than muddled in the shaker or used as a garnish, the leaves are shaken whole with the other ingredients and caught by the strainer. This produces a gentler mint character than muddling — more aromatic than aggressive, woven into the drink rather than dominating it.""",
        "notes": """Creator's spec per Punch, January 2025. A botanical gin with genuine herbal presence works better here than a neutral or citrus-forward gin — the mint needs something to harmonize with. Beefeater, Tanqueray, or Sipsmith are appropriate. The double-strain or tightly gated Hawthorne is important — mint leaf fragments in a coupe are not pleasant. Aperol's lower proof (11% ABV) means this is the lowest-alcohol drink in the Bug family; factor that in if you are building a round.""",
        "riffs": """Substitute Campari for Aperol for a significantly more bitter, less sweet version — closer to a gin Mosquito. Add a half ounce of green Chartreuse and reduce Aperol to 0.5 oz to push the drink back toward its Last Word ancestor. Basil shaken in place of mint produces a more savory, anise-adjacent variation worth trying.""",
        "why_it_matters": "The Maverick closes the circle — a gin drink descended from a gin drink, routed through bourbon and mezcal along the way. The most approachable entry point into the Bug family, and a demonstration that the equal-parts template is as comfortable with lightness as with complexity.",
        "curiosity_hook": "Shaking mint rather than muddling it is a technique choice with real consequences. Muddling ruptures the cell walls and releases chlorophyll along with the aromatics — producing a greener, more aggressive, sometimes bitter mint character. Shaking whole leaves extracts aromatics more gently. Neither is wrong; they produce different results. This drink needs the gentle version.",
        "source_urls": SOURCE_URL,
    },
    {
        "recipe_id": "lantern_fly",
        "name": "Lantern Fly",
        "creator_id": "jojo_colona_attaboy",
        "origin": "Jojo Colona, Attaboy, New York City",
        "base_spirit": "Rum",
        "tags": "Sour/Tart, Bitter, Fruity, Boozy, Spicy",
        "sliders": dict(boozy=6, sweet=3, sour=6, bitter=7, fruity=5, herbal=4, smoky=0, spicy=4, rich=2),
        "related_recipe_ids": "mosquito,paper_plane,last_word",
        "ingredients": [
            ("doctor_bird_jamaican_pot_still_rum", "Jamaican pot still rum (Doctor Bird recommended)", "0.75", "oz"),
            ("fresh_lemon_juice", "Fresh lemon juice", "0.75", "oz"),
            ("sweetened_ginger_juice", "Sweetened ginger juice", "0.75", "oz"),
            ("amargo_vallet", "Amargo-Vallet amaro", "0.75", "oz"),
        ],
        "instruction": (
            "Combine all ingredients in a cocktail shaker with ice. Shake until well chilled. "
            "Strain into a chilled coupe. Garnish with candied ginger."
        ),
        "garnish": ("Candied ginger", None, "coupe rim or pick"),
        "story": """Jojo Colona created the Lantern Fly at Attaboy's New York location, working with two ingredients that are unusual choices within the Bug family and that produce, together, the most distinctly funky drink in the series.

Jamaican pot still rum is not a background ingredient. Made in traditional pot stills rather than the continuous column stills that produce lighter Caribbean rums, pot still Jamaican rum carries an intense, funky, almost overripe fruit quality — esters and congeners that make themselves known in any drink they inhabit. Doctor Bird, the rum Punch specifies for this recipe, is a Worthy Park expression named for the Jamaican national bird, made entirely from pot still distillate. It is not a subtle rum.

Amargo-Vallet is equally specific. The Mexican digestif amaro, made since 1881 in Mexico City and named for its French founder Adolphe Vallet, is built on a base of angostura bark — making it a relative of Angostura bitters in botanical character, with the flavor profile of a bitter liqueur rather than a concentrate. It is drier, more herbal, and less sweet than Campari, and its angostura-forward bitterness creates a different kind of tension with the ginger than Campari's more fruit-and-citrus bitterness.

The combination — funky pot still rum, angostura-forward amaro, ginger heat, lemon acid — is aggressive and layered in a way that rewards attention.""",
        "notes": """Creator's spec per Punch, January 2025. Jamaican pot still rum is not interchangeable with standard Jamaican rum or blended Caribbean rum — the funk is structural. Worthy Park, Hampden, or Doctor Bird are the appropriate choices. Amargo-Vallet has limited distribution; check specialty spirits retailers. In a pinch, Angostura bitters at 0.5 oz plus 0.25 oz Campari approximates the botanical character at some cost to the amaro's depth and sweetness contribution — adjust the ginger juice accordingly.""",
        "riffs": """Substitute Campari for Amargo-Vallet for a sweeter, more familiar bitter profile that tames the funk of the pot still rum rather than amplifying it. Smith & Cross, the Navy-strength Jamaican pot still rum, in place of Doctor Bird increases the proof significantly and pushes the funk further — the drink becomes considerably more demanding. Combine equal parts pot still rum and aged agricole rhum for a split-base variation with additional grassy, vegetal complexity.""",
        "why_it_matters": "The most ingredient-specific drink in the Bug family, and the one most likely to reward the curious drinker who seeks it out. Amargo-Vallet is genuinely obscure outside of cocktail enthusiast circles; Jamaican pot still rum is underappreciated even among rum drinkers. The Lantern Fly puts both in a context where they cannot hide.",
        "curiosity_hook": "Amargo-Vallet is a case study in how a drink's bitter component determines its character as much as the base spirit does. Swap Amargo-Vallet for Campari in this recipe and you have a different drink — same structure, same proportions, genuinely different experience. The bitter element in an equal-parts Daisy is not interchangeable. Choose it as carefully as you choose the spirit.",
        "source_urls": "https://punchdrink.com/recipes/lantern-fly/\n" + SOURCE_URL,
    },
]


def append_row(ws, values: list):
    ws.append(values)


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)

    # Taxonomy extension for Daisy — Equal Parts
    families = wb["Families"]
    existing_family_ids = {
        row[0].value for row in families.iter_rows(min_row=2) if row[0].value
    }
    if "daisy_equal_parts" not in existing_family_ids:
        append_row(
            families,
            ["daisy_equal_parts", "Daisy — Equal Parts", "subfamily", "sour", "draft_needs_review"],
        )

    # Pack
    packs = wb["Packs"]
    pack_ids = {row[0].value for row in packs.iter_rows(min_row=2) if row[0].value}
    if PACK_ID not in pack_ids:
        append_row(
            packs,
            [
                PACK_ID,
                PACK_NAME,
                1,
                "draft_needs_review",
                len(COCKTAILS),
                "liquid_alchemy_attaboy_bug_cocktails.txt",
                "2026-06-04",
                "Attaboy Bug Cocktail family: Mosquito, Praying Mantis, Maverick, Lantern Fly.",
            ],
        )

    # Creators
    creators = wb["Creators"]
    existing_creator_ids = {
        row[0].value for row in creators.iter_rows(min_row=2) if row[0].value
    }
    for creator_id, data in CREATORS.items():
        if creator_id not in existing_creator_ids:
            append_row(
                creators,
                [creator_id, data["name"], data["bio"], data["source_urls"], "draft_needs_review"],
            )

    # Ingredients
    ingredients = wb["Ingredients"]
    existing_ingredient_ids = {
        row[0].value for row in ingredients.iter_rows(min_row=2) if row[0].value
    }
    for ingredient_id, name, category, flavor_notes in NEW_INGREDIENTS:
        if ingredient_id not in existing_ingredient_ids:
            append_row(
                ingredients,
                [ingredient_id, name, None, flavor_notes, category, "draft_needs_review"],
            )

    # Craft item (single shared prep)
    craft_items = wb["Craft_Items"]
    craft_ids = {row[0].value for row in craft_items.iter_rows(min_row=2) if row[0].value}
    if CRAFT_ITEM_ID not in craft_ids:
        append_row(
            craft_items,
            [
                CRAFT_ITEM_ID,
                "Sweetened Ginger Juice",
                "homemade_ingredient",
                "juice",
                "ginger_juice",
                "sam_ross_attaboy",
                "Easy",
                "Varies by batch; enough for multiple cocktails",
                "Refrigerate up to 2 days",
                "PUNCH / Attaboy",
                SOURCE_URL,
                "Fresh-pressed ginger intensity is non-negotiable for the Bug family.",
                "Shared house prep across the Attaboy Bug Cocktail family.",
                CRAFT_GINGER_TEXT,
                "draft_needs_review",
                PACK_ID,
                "NO - Craft is not in current app import/export",
            ],
        )

    craft_ingredients = wb["Craft_Ingredients"]
    if CRAFT_ITEM_ID not in {
        row[0].value for row in craft_ingredients.iter_rows(min_row=2) if row[0].value
    }:
        append_row(
            craft_ingredients,
            [CRAFT_ITEM_ID, 1, "Fresh ginger (skin on)", "4", "parts", "juiced", "By weight relative to sugar"],
        )
        append_row(
            craft_ingredients,
            [CRAFT_ITEM_ID, 2, "Granulated sugar", "3", "parts", None, "By weight; stir until dissolved"],
        )

    craft_steps = wb["Craft_Steps"]
    if CRAFT_ITEM_ID not in {
        row[0].value for row in craft_steps.iter_rows(min_row=2) if row[0].value
    }:
        append_row(
            craft_steps,
            [
                CRAFT_ITEM_ID,
                1,
                "Juice fresh ginger skin-on to extract fresh ginger juice.",
                None,
                None,
            ],
        )
        append_row(
            craft_steps,
            [
                CRAFT_ITEM_ID,
                2,
                "Mix 4 parts ginger juice to 3 parts granulated sugar by weight, stirring until sugar fully dissolves.",
                None,
                "Keeps refrigerated up to 2 days.",
            ],
        )

    craft_related = wb["Craft_Related_Recipes"]
    existing_related = {
        (row[0].value, row[1].value)
        for row in craft_related.iter_rows(min_row=2)
        if row[0].value and row[1].value
    }

    recipes_ws = wb["Recipes"]
    recipe_headers = [cell.value for cell in next(recipes_ws.iter_rows(min_row=1, max_row=1))]
    recipe_idx = {name: i for i, name in enumerate(recipe_headers)}

    ings_ws = wb["Recipe_Ingredients"]
    steps_ws = wb["Steps"]
    lore_ws = wb["Lore"]
    garnishes_ws = wb["Garnishes"]
    export_ws = wb["App_JSON_Export"]
    export_headers = [cell.value for cell in next(export_ws.iter_rows(min_row=1, max_row=1))]

    existing_recipe_ids = {
        row[recipe_idx["recipe_id"]].value
        for row in recipes_ws.iter_rows(min_row=2)
        if row[recipe_idx["recipe_id"]].value
    }

    base_added_at = 1780281600000
    for offset, cocktail in enumerate(COCKTAILS):
        recipe_id = cocktail["recipe_id"]
        if recipe_id in existing_recipe_ids:
            raise SystemExit(f"Recipe already exists: {recipe_id}")

        append_row(
            recipes_ws,
            [
                recipe_id,
                cocktail["name"],
                "sour",
                "Sour",
                "daisy_equal_parts",
                "Daisy — Equal Parts",
                cocktail["base_spirit"],
                "coupe",
                "Coupe",
                "Up",
                "Medium" if recipe_id in {"praying_mantis", "lantern_fly"} else "Easy",
                "After Dinner" if recipe_id != "maverick" else "Aperitivo",
                "Year-Round",
                cocktail["tags"],
                0,
                base_added_at + offset,
                cocktail["sliders"]["boozy"],
                cocktail["sliders"]["sweet"],
                cocktail["sliders"]["sour"],
                cocktail["sliders"]["bitter"],
                cocktail["sliders"]["fruity"],
                cocktail["sliders"]["herbal"],
                cocktail["sliders"]["smoky"],
                cocktail["sliders"]["spicy"],
                cocktail["sliders"]["rich"],
                PACK_ID,
                1,
                "draft_needs_review",
                cocktail["curiosity_hook"],
                cocktail["why_it_matters"],
                cocktail["related_recipe_ids"],
                "FALSE",
                None,
                False,
                json.dumps([CRAFT_ITEM_ID]),
                SOURCE_URL,
            ],
        )

        for sort_order, ingredient in enumerate(cocktail["ingredients"], start=1):
            ingredient_id, name, amount, unit = ingredient[:4]
            note = ingredient[4] if len(ingredient) > 4 else None
            append_row(
                ings_ws,
                [recipe_id, sort_order, ingredient_id, name, amount, unit, None, note, "FALSE", None],
            )

        append_row(
            steps_ws,
            [recipe_id, 1, cocktail["instruction"], "attaboy_bug_family_import"],
        )

        append_row(
            lore_ws,
            [
                recipe_id,
                cocktail["story"],
                cocktail["origin"],
                cocktail["why_it_matters"],
                cocktail["curiosity_hook"],
                cocktail["notes"],
                cocktail["riffs"],
                None,
                cocktail["creator_id"],
                cocktail["source_urls"],
                SOURCE_NOTES,
                "needs_review",
            ],
        )

        garnish_name, garnish_prep, garnish_placement = cocktail["garnish"]
        append_row(
            garnishes_ws,
            [recipe_id, 1, garnish_name, garnish_prep, garnish_placement, "attaboy_bug_family_import"],
        )

        kebab_id = recipe_id.replace("_", "-")
        export_row = {header: None for header in export_headers}
        export_row.update(
            {
                "id": kebab_id,
                "name": cocktail["name"],
                "export_ready": "YES",
                "export_notes": "Images intentionally blank for current app bulk import; use asset pipeline later.",
                "pack_id": PACK_ID,
                "status": "draft_needs_review",
            }
        )
        append_row(export_ws, [export_row.get(header) for header in export_headers])

        for related_kebab in [recipe_id.replace("_", "-")]:
            key = (CRAFT_ITEM_ID, related_kebab)
            if key not in existing_related:
                append_row(
                    craft_related,
                    [
                        CRAFT_ITEM_ID,
                        related_kebab,
                        "required_prep",
                        f"Required house prep for {cocktail['name']}.",
                    ],
                )
                existing_related.add(key)

    wb.save(WORKBOOK)
    print(f"Imported {len(COCKTAILS)} cocktails into {WORKBOOK}")


if __name__ == "__main__":
    main()
